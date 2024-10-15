import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import psycopg2
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time

result_checkbox = ""
# massive = [[0] * row_T1 for i in range(row_T2)]

def open_excel_file_T1():
    # Открываем файловый диалог для выбора файла Excel
    file_path_T1 = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path_T1:
        # Открываем файл Excel с помощью openpyxl
        wb2 = load_workbook(file_path_T1)
        # Выводим сообщение о успешном открытии файла
        label.config(text=f"Файл {file_path_T1} открыт успешно!")

        # Выполняем основной код
        process_excel_file_T1(wb2)

def open_excel_file_T2():
    # Открываем файловый диалог для выбора файла Excel
    file_path_T2 = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path_T2:
        # Открываем файл Excel с помощью openpyxl
        wb3 = load_workbook(file_path_T2)
        # Выводим сообщение о успешном открытии файла
        label.config(text=f"Файл {file_path_T2} открыт успешно!")

        # Выполняем основной код
        process_excel_file_T2(wb3)


def process_excel_file_T1(wb2):
    # Получаем начальное время
    st = time.time()

    print(wb2.sheetnames)
    sheet_obj = wb2.active
    print("Size:" + str(sheet_obj.max_column) + " X " + str(sheet_obj.max_row))
    count_column = sheet_obj.max_column
    count_row = sheet_obj.max_row
    cell_obj = sheet_obj.cell(row=1, column=1)

    color_arr = ["00000000", "1", "2", "FF8DB4E2", "0", "FFDA9694", "5", 0, 0, 0, 0, 0]
    print(color_arr[5])

    fooo = [[0] * count_column for i in range(count_row)]
    fooo_color = [[0] * count_column for i in range(count_row)]

    for key, value2 in enumerate(fooo):
        for key1, value1 in enumerate(fooo[key]):
            fooo[key][key1] = str(sheet_obj.cell(row=key + 1, column=key1 + 1).value).strip()
            colorr = str(sheet_obj.cell(row=key + 1, column=key1 + 1).fill.fgColor.rgb)
            if colorr[0] == "V":
                colorr = int(sheet_obj.cell(row=key + 1, column=key1 + 1).fill.fgColor.theme)
                if colorr > 11:
                    colorr = 0
                colorr = color_arr[colorr]
            fooo_color[key][key1] = colorr
    print(fooo[0][3])

    # Подключение к базе данных
    conn = psycopg2.connect(
        dbname="vibory",
        user="elverona",
        password="qwerty",
        host="localhost",
        port="5432"
    )

    # Создание курсора для выполнения операций с базой данных
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS T1")
    command_T1 = ''
    for i in range(count_column):
        command_T1 += 'column' + str(i) + ' VARCHAR(400),'
    command_T1 += 'columnError VARCHAR(100)'
    command_init_T1 = "CREATE TABLE T1 (" + command_T1 + ")"
    cur.execute(command_init_T1)
    insert_command_T1 = "INSERT INTO T1 VALUES (" + ",".join(["%s"] * count_column) + ", %s)"
    for row in fooo:
        cur.execute(insert_command_T1, row + [None])

    conn.commit()
    conn.close()

    workbook = openpyxl.Workbook()

    # Create a new sheet
    sheet = workbook.active

    # Write the array data to the sheet
    for row_index, row_data in enumerate(fooo_color):
        for column_index, color in enumerate(row_data):
            cell = sheet.cell(row=row_index + 1, column=column_index + 1)
            cell.value = fooo[row_index][column_index]
            if isinstance(color, int):
                color = hex(color)[2:].upper()
            if color != "00000000" and color != "0":
                fill = PatternFill(start_color =color, end_color=color, fill_type="solid")
                cell.fill = fill

    # Save the workbook to a file
    workbook.save('C:/1/2.xlsx')

    print("Excel file 'C:/1/2.xlsx' created successfully.")


def process_excel_file_T2(wb3):
    # Получаем начальное время
    st = time.time()

    print(wb3.sheetnames)
    sheet_obj = wb3.active
    print("Size:" + str(sheet_obj.max_column) + " X " + str(sheet_obj.max_row))
    count_column = sheet_obj.max_column
    count_row = sheet_obj.max_row
    cell_obj = sheet_obj.cell(row=1, column=1)

    color_arr = ["00000000", "1", "2", "FF8DB4E2", "0", "FFDA9694", "5", 0, 0, 0, 0, 0]
    print(color_arr[5])

    fooo = [[0] * count_column for i in range(count_row)]
    fooo_color = [[0] * count_column for i in range(count_row)]

    for key, value2 in enumerate(fooo):
        for key1, value1 in enumerate(fooo[key]):
            fooo[key][key1] = str(sheet_obj.cell(row=key + 1, column=key1 + 1).value).strip()
            colorr = str(sheet_obj.cell(row=key + 1, column=key1 + 1).fill.fgColor.rgb)
            if colorr[0] == "V":
                colorr = int(sheet_obj.cell(row=key + 1, column=key1 + 1).fill.fgColor.theme)
                if colorr > 11:
                    colorr = 0
                colorr = color_arr[colorr]
            fooo_color[key][key1] = colorr
    print(fooo[0][3])

    # Подключение к базе данных
    conn = psycopg2.connect(
        dbname="vibory",
        user="elverona",
        password="qwerty",
        host="localhost",
        port="5432"
    )

    # Создание курсора для выполнения операций с базой данных
    cur = conn.cursor()

    cur.execute("DROP TABLE IF EXISTS T2")
    command_T2 = ''
    for i in range(count_column):
        command_T2 += 'column' + str(i) + ' VARCHAR(300),'
    command_T2 += 'columnError VARCHAR(100)'
    command_init_T2 = "CREATE TABLE T2 (" + command_T2 + ")"
    cur.execute(command_init_T2)
    insert_command_T2 = "INSERT INTO T2 VALUES (" + ",".join(["%s"] * count_column) + ", %s)"
    for row in fooo:
        cur.execute(insert_command_T2, row + [None])

    conn.commit()
    conn.close()

    workbook = openpyxl.Workbook()

    # Create a new sheet
    sheet = workbook.active

    # Write the array data to the sheet
    for row_index, row_data in enumerate(fooo_color):
        for column_index, color in enumerate(row_data):
            cell = sheet.cell(row=row_index + 1, column=column_index + 1)
            cell.value = fooo[row_index][column_index]
            if isinstance(color, int):
                color = hex(color)[2:].upper()
            if color != "00000000" and color != "0":
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.fill = fill

    # Save the workbook to a file
    workbook.save('C:/1/3.xlsx')

    print("Excel file 'C:/1/3.xlsx' created successfully.")

    # получаем время завершения
    et = time.time()

    # считаем время исполнения
    elapsed_time = et - st
    print('Время исполнения:', elapsed_time, 'секунд')


def checkbox_changed(checkbox_var, row, column, ii):
    global result_checkbox
    if checkbox_var.get():
        print(f"Чекбокс в строке {row} и столбце {column} активирован")
        if ii:
            result_checkbox = str(int(ii) + (row % 10) + (column % 10))
        else:
            result_checkbox = str((row % 10) + (column % 10))
    else:
        print(f"Чекбокс в строке {row} и столбце {column} деактивирован")
        result_checkbox = ""

def action():
    s = ""
    p = ""
    k = ""
    for i in range(0, len(massive)):
        for i2 in range(0, len(massive[i])):
            if (i>0) and (i2>0):
                if massive[i][i2] > 0:
                    # print(massive[i][0], 'связь', massive[i][i2], 'С', massive[0][i2])
                    s = s + massive[i][0] + ">" + massive[0][i2] + "\n" if massive[i][i2] & 1 else s
                    p = p + massive[i][0] + ">" + massive[0][i2] + "\n" if (massive[i][i2] >> 1) & 1 else p
                    k = k + massive[i][0] + ">" + massive[0][i2] + "\n" if (massive[i][i2] >> 2) & 1 else k
    print("Связь:\n", s, "Приоритет:\n", p, "Ключ:\n", k)
    # if s:
        

#     print(massive[i][i2], end=',')
        # print()
def checkbox_changed(checkbox_var, row, column):
    global result_checkbox, b, a
    b = row & 7
    a = row >> 3
    if checkbox_var.get():
        # print(f"Чекбокс в строке {row} и столбце {column} активирован")
        massive[a][column] = massive[a][column] | b
    else:
        # print(f"Чекбокс в строке {row} и столбце {column} деактивирован")
        massive[a][column] = massive[a][column] & (~b)
    # print("S,P,K\n", (massive[1][1] & 1), ((massive[1][1] >> 1) & 1), ((massive[1][1] >> 2) & 1))
# Карта действий:
# 1. вывести наименование столбцов и строк с номером строки или стобца

def create_T3():
    # Подключение к базе данных
    conn = psycopg2.connect(
        dbname="vibory",
        user="elverona",
        password="qwerty",
        host="localhost",
        port="5432"
    )

    # Создание курсора для выполнения операций с базой данных
    cur = conn.cursor()

    # Получение первой строки из таблицы T1
    cur.execute("SELECT * FROM T1 LIMIT 1")
    row_T1 = cur.fetchone()

    # Получение первой строки из таблицы T2
    cur.execute("SELECT * FROM T2 LIMIT 1")
    row_T2 = cur.fetchone()

    global massive
    massive = [[0] * (len(row_T1) + 1) for i in range((len(row_T2)) + 1)]
    for p in enumerate(row_T1):
        massive[0][p[0] + 1] = p[1]
    for o in enumerate(row_T2):
        massive[o[0] + 1][0] = o[1]
    # print(massive, end="\n")
    # for i in range(0, len(massive)):
    #     for i2 in range(0, len(massive[i])):
    #         print(massive[i][i2], end=',')
    #     print()



    # Создание нового окна для таблицы
    table_window = tk.Toplevel(root)
    table_window.title("Таблица T3")

    # Создание таблицы в новом окне
    table_frame = tk.Frame(table_window)
    table_frame.pack(fill="both", expand=True)

    # Создание стиля для полосы прокрутки
    style = ttk.Style()
    style.configure("Horizontal.TScrollbar", thumbcolor='gray50')
    style.configure("Vertical.TScrollbar", thumbcolor='gray50')

    # Создание горизонтальной полосы прокрутки
    hscrollbar = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, style="Horizontal.TScrollbar")
    hscrollbar.pack(side=tk.BOTTOM, fill=tk.X)

    # Создание вертикальной полосы прокрутки
    vscrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, style="Vertical.TScrollbar")
    vscrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Создание канвы для таблицы
    canvas = tk.Canvas(table_frame, width=400, height=200)
    canvas.pack(side=tk.LEFT, fill="both", expand=True)

    # Создание фрейма для содержимого таблицы
    table_content_frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=table_content_frame, anchor="nw")

    # Создание заголовков таблицы
    for i, value in enumerate(row_T1):
        label = tk.Label(table_content_frame, text=str(value), highlightthickness=1, highlightbackground="gray")
        label.grid(row=0, column=i + 1, padx=2, pady=2)

    # Создание строк таблицы

    for i, value in enumerate(row_T2):
        label = tk.Label(table_content_frame, text=str(value), highlightthickness=1, highlightbackground="gray")
        label.grid(row=i + 1, column=0, padx=2, pady=2)
        for j in range(len(row_T1)):
            checkbox_frame = tk.Frame(table_content_frame, highlightthickness=1, highlightbackground="gray")
            checkbox_frame.grid(row=i + 1, column=j + 1, padx=2, pady=2)
            checkbox_vars = [tk.IntVar() for _ in range(3)]
            for k, checkbox_var in enumerate(checkbox_vars):
                checkbox = tk.Checkbutton(checkbox_frame, variable=checkbox_var,
                                          command=lambda var=checkbox_var, row=(((i+1)<<3) + (2**k)), column=(j + 1): checkbox_changed(
                                              var, row, column))
                checkbox.pack(side=tk.LEFT)

    # Конфигурация полосы прокрутки
    hscrollbar.config(command=canvas.xview)
    vscrollbar.config(command=canvas.yview)
    canvas.config(xscrollcommand=hscrollbar.set, yscrollcommand=vscrollbar.set)



    # Создаем кнопку выполнения действия
    button_execute = tk.Button(table_window, text="Выполнить действие", command=lambda: action())
    button_execute.pack()

    # Закрытие соединения с базой данных
    conn.close()

    # Вывод сообщения о создании таблицы
    label = tk.Label(table_window, text="Таблица T3 создана успешно!")
    label.pack()


def create_button(frame, text, command):
    # Создаем кнопку
    button = tk.Button(frame, text=text, command=command)
    button.pack(side=tk.LEFT, padx=5, pady=5)
    return button


# Создаем главное окно
root = tk.Tk()
root.title("Открытие документов Excel")

# Создаем фрейм для кнопок
frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

# Создаем кнопки
button_open_T1 = create_button(frame, "Открыть файл Excel для Т1", open_excel_file_T1)
button_open_T2 = create_button(frame, "Открыть файл Excel для Т2", open_excel_file_T2)
button_create_T3 = create_button(frame, "Создать T3", create_T3)
# button_create_T 4 = create_button(frame, "Создать T4", create_T4)
button_exit = create_button(frame, "Выход", root.destroy)

# Создаем метку для вывода сообщений
label = tk.Label(root, text="")
label.pack(padx=10, pady=10)

# Запускаем главное окно
root.mainloop()